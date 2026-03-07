"""
Deterministic Spreadsheet Ingestion & Structured Retrieval Cache.

Goals:
1. Preserve complex layouts (tabular, key-value, hierarchical).
2. Build deterministic JSON records for exact lookup.
3. Embed row-level records for semantic fallback.
"""

import datetime
import json
import os
import re
import time
import warnings
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from langchain_core.documents import Document
from langchain_qdrant import QdrantVectorStore
from qdrant_client.http import models

from database import get_qdrant_client
from embeddings_provider import get_embeddings

warnings.filterwarnings("ignore")

# ── Configuration ──────────────────────────────────────────────────
COLLECTION_NAME = "local_documents"
STRUCTURED_CACHE_DIR = os.path.join("data", "structured_cache")
os.makedirs(STRUCTURED_CACHE_DIR, exist_ok=True)

# Global cache for instant structured lookup
EXCEL_CACHE: Dict[str, Any] = {}


# ── Helpers ────────────────────────────────────────────────────────
def _to_snake_case(text: str) -> str:
    s = str(text).strip()
    s = re.sub(r"([a-z0-9])([A-Z])", r"\1_\2", s)
    s = re.sub(r"[^a-zA-Z0-9]", "_", s).lower()
    return re.sub(r"_+", "_", s).strip("_")


def _clean_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    return str(value).strip()


def _non_empty_values(row: List[str]) -> List[str]:
    return [v for v in row if v]


def _is_numeric(val: Any) -> bool:
    try:
        v = _clean_cell(val)
        if not v:
            return False
        float(v.replace(",", "").replace("%", ""))
        return True
    except Exception:
        return False


def _looks_like_header(row: List[Any]) -> bool:
    non_empty = [v for v in row if _clean_cell(v)]
    if len(non_empty) < 2:
        return False
    string_count = sum(1 for v in non_empty if not _is_numeric(v) and len(_clean_cell(v)) > 1)
    return string_count >= len(non_empty) * 0.7


def _header_score(row: List[str], row_idx: int) -> float:
    non_empty = _non_empty_values(row)
    if len(non_empty) < 2:
        return 0.0
    unique_ratio = len(set(non_empty)) / max(1, len(non_empty))
    alpha_count = sum(1 for v in non_empty if re.search(r"[A-Za-z]", v))
    length_bonus = len(non_empty) * 0.1
    index_penalty = row_idx * 0.05
    return unique_ratio + length_bonus + alpha_count - index_penalty


def _dedupe_headers(headers: List[str]) -> List[str]:
    seen: Dict[str, int] = {}
    result = []
    for i, h in enumerate(headers):
        base = _to_snake_case(h) or f"col_{i + 1}"
        count = seen.get(base, 0)
        seen[base] = count + 1
        result.append(base if count == 0 else f"{base}_{count + 1}")
    return result


def _detect_layout(df: pd.DataFrame) -> str:
    if df.empty:
        return "unknown"

    # Improved tabular detection: if we see columns with headers like "Plan" or "Category"
    potential_headers = []
    for i in range(min(15, len(df))):
        row = df.iloc[i].tolist()
        if _looks_like_header(row):
            header_found = True
            potential_headers = row
            break

    # If many columns have headers containing "plan", it is likely tabular
    if any("plan" in str(h).lower() for h in potential_headers):
        return "tabular"

    non_empty_cols = [c for c in df.columns if not df[c].astype(str).str.strip().eq("").all()]
    if len(non_empty_cols) == 2:
        col1 = df[non_empty_cols[0]].astype(str).str.strip()
        if col1.nunique() > len(df) * 0.4:
            return "key_value"

    if header_found:
        return "tabular"
    return "hierarchical"


def _trim_grid(grid: List[List[str]]) -> List[List[str]]:
    if not grid:
        return grid

    # Remove trailing empty rows.
    while grid and not any(cell for cell in grid[-1]):
        grid.pop()
    if not grid:
        return grid

    # Remove trailing empty columns.
    max_cols = max(len(r) for r in grid)
    last_non_empty_col = -1
    for c in range(max_cols):
        for r in range(len(grid)):
            if c < len(grid[r]) and grid[r][c]:
                last_non_empty_col = c
                break
    if last_non_empty_col == -1:
        return []

    return [row[: last_non_empty_col + 1] for row in grid]


def _read_xlsx_with_layout(file_path: str) -> Dict[str, pd.DataFrame]:
    from openpyxl import load_workbook

    wb = load_workbook(file_path, data_only=True, read_only=False)
    frames: Dict[str, pd.DataFrame] = {}
    for ws in wb.worksheets:
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        if max_row == 0 or max_col == 0:
            frames[ws.title] = pd.DataFrame()
            continue

        grid = [["" for _ in range(max_col)] for _ in range(max_row)]
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                grid[r - 1][c - 1] = _clean_cell(ws.cell(row=r, column=c).value)

        # Propagate merged-cell values so layouts are not lost.
        for merged in ws.merged_cells.ranges:
            value = _clean_cell(ws.cell(merged.min_row, merged.min_col).value)
            for r in range(merged.min_row, merged.max_row + 1):
                for c in range(merged.min_col, merged.max_col + 1):
                    grid[r - 1][c - 1] = value

        trimmed = _trim_grid(grid)
        frames[ws.title] = pd.DataFrame(trimmed) if trimmed else pd.DataFrame()

    return frames


def _read_sheet_frames(file_path: str, ext: str) -> Dict[str, pd.DataFrame]:
    if ext in (".xlsx", ".xlsm"):
        return _read_xlsx_with_layout(file_path)

    if ext == ".xls":
        xl = pd.ExcelFile(file_path, engine="xlrd")
        return {sheet: xl.parse(sheet, header=None).fillna("") for sheet in xl.sheet_names}

    if ext == ".csv":
        return {"CSV_DATA": pd.read_csv(file_path, header=None).fillna("")}

    return {}


def _find_header_index(rows: List[List[str]]) -> Optional[int]:
    best_idx: Optional[int] = None
    best_score = 0.0
    for idx, row in enumerate(rows[:14]):
        if not _looks_like_header(row):
            continue
        score = _header_score(row, idx)
        if score > best_score:
            best_score = score
            best_idx = idx
    if best_score < 1.0:
        return None
    return best_idx


def _sanitize_filename(name: str) -> str:
    safe = re.sub(r"[^a-zA-Z0-9._-]+", "_", name).strip("_")
    return safe or "structured_cache.json"


def _cache_path(file_name: str, folder_name: str) -> str:
    safe_file = _sanitize_filename(file_name)
    safe_folder = _sanitize_filename(folder_name)
    return os.path.join(STRUCTURED_CACHE_DIR, f"{safe_folder}__{safe_file}.json")


def _persist_structured_json(file_name: str, folder_name: str, payload: Dict[str, Any]) -> str:
    path = _cache_path(file_name, folder_name)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False)
    return path


def load_structured_payloads(folder_name: str = "All") -> List[Dict[str, Any]]:
    payloads: List[Dict[str, Any]] = []
    seen = set()

    for payload in EXCEL_CACHE.values():
        if not isinstance(payload, dict):
            continue
        if folder_name != "All" and payload.get("folder") != folder_name:
            pf = str(payload.get("folder", "")).lower()
            fn = folder_name.lower()
            if not (pf in fn or fn in pf):
                continue
        cache_key = (payload.get("folder"), payload.get("file_name"))
        payloads.append(payload)
        seen.add(cache_key)

    if not os.path.isdir(STRUCTURED_CACHE_DIR):
        return payloads

    for file_name in os.listdir(STRUCTURED_CACHE_DIR):
        if not file_name.endswith(".json"):
            continue
        path = os.path.join(STRUCTURED_CACHE_DIR, file_name)
        try:
            with open(path, "r", encoding="utf-8") as f:
                payload = json.load(f)
            if not isinstance(payload, dict):
                continue
            if folder_name != "All" and payload.get("folder") != folder_name:
                pf = str(payload.get("folder", "")).lower()
                fn = folder_name.lower()
                if not (pf in fn or fn in pf):
                    continue
            cache_key = (payload.get("folder"), payload.get("file_name"))
            if cache_key in seen:
                continue
            payloads.append(payload)
            seen.add(cache_key)
        except Exception:
            continue

    return payloads


def delete_structured_cache(folder_name: str, file_name: Optional[str] = None) -> Dict[str, int]:
    """
    Remove structured cache entries for a folder, optionally scoped to one file.
    Returns counts for in-memory and disk deletions.
    """
    removed_memory = 0
    removed_disk = 0

    # In-memory cache cleanup.
    for cache_key, payload in list(EXCEL_CACHE.items()):
        if not isinstance(payload, dict):
            continue
        if payload.get("folder") != folder_name:
            continue
        if file_name and payload.get("file_name") != file_name:
            continue
        EXCEL_CACHE.pop(cache_key, None)
        removed_memory += 1

    # Disk cache cleanup.
    if os.path.isdir(STRUCTURED_CACHE_DIR):
        for entry in os.listdir(STRUCTURED_CACHE_DIR):
            if not entry.endswith(".json"):
                continue
            path = os.path.join(STRUCTURED_CACHE_DIR, entry)
            try:
                with open(path, "r", encoding="utf-8") as f:
                    payload = json.load(f)
            except Exception:
                continue
            if not isinstance(payload, dict):
                continue
            if payload.get("folder") != folder_name:
                continue
            if file_name and payload.get("file_name") != file_name:
                continue
            try:
                os.remove(path)
                removed_disk += 1
            except Exception:
                continue

    return {"removed_memory": removed_memory, "removed_disk": removed_disk}


USE_LLM_FLATTENER = True
OLLAMA_BASE_URL = os.getenv("OLLAMA_BASE_URL", "http://localhost:11434")
OLLAMA_FLATTEN_MODEL = "qwen2.5:14b"

def _llm_flatten_sheet_to_json(df: pd.DataFrame, sheet_name: str) -> Optional[List[Dict[str, str]]]:
    """
    Uses an LLM (Cloud or Local) to take a chaotic, merged-cell Excel sheet
    and flatten it into a perfect, uniform JSON array of key-value pairs.
    """
    if not USE_LLM_FLATTENER or df.empty:
        return None

    try:
        import requests
        csv_data = df.to_csv(index=False)
        # Limit enormous sheets to prevent memory explosion in the fast prompt
        if len(csv_data) > 30000:
            csv_data = csv_data[:30000] + "\n...[TRUNCATED]"

        prompt = f"""You are a master Data Engineer. Below is the raw CSV data of an Excel sheet named '{sheet_name}'.
The sheet contains merged cells, nested headers, and chaotic structure.
Your job is to read it, understand the visual layout and hierarchy, and FLATTEN it into a uniform JSON array.

RULES:
1. Output ONLY valid JSON starting with `[` and ending with `]`. No markdown tags like ```json.
2. Each object in the array should represent a distinct fact or row constraint.
3. Determine the correct 'Entity' (e.g. Employee, Dependant, etc.) and 'Category' (e.g. Outpatient, Inpatient) and group them properly as keys.
4. DO NOT output any text outside of the JSON array.
5. If the table is empty or garbage, return [].

RAW CSV DATA:
{csv_data}
"""
        print(f"      [LLM-FLATTEN] Calling {OLLAMA_FLATTEN_MODEL} to flatten sheet '{sheet_name}'...")
        payload = {
            "model": OLLAMA_FLATTEN_MODEL,
            "prompt": prompt,
            "stream": False,
            "options": {"temperature": 0.0}
        }
        resp = requests.post(f"{OLLAMA_BASE_URL}/api/generate", json=payload, timeout=120)
        if resp.status_code == 200:
            raw_text = resp.json().get("response", "").strip()
            # Clean up markdown if model outputs it
            if raw_text.startswith("```json"):
                raw_text = raw_text[7:]
            if raw_text.startswith("```"):
                raw_text = raw_text[3:]
            if raw_text.endswith("```"):
                raw_text = raw_text[:-3]
            try:
                parsed = json.loads(raw_text.strip())
                if isinstance(parsed, list):
                    print(f"      [LLM-FLATTEN] Success! Extracted {len(parsed)} clean records.")
                    return parsed
            except Exception as e:
                print(f"      [LLM-FLATTEN] JSON Parse Error: {e}")
                return None
    except Exception as e:
        print(f"      [LLM-FLATTEN] Connection/Timeout Error: {e}")
        return None
    return None



def _df_to_markdown_table(df: pd.DataFrame) -> str:
    if df.empty: return ""
    # Safe markdown table generation without relying on tabulate
    headers = " | ".join(str(c).replace("\n", " ").replace("|", "-") for c in df.columns)
    separator = " | ".join("---" for _ in df.columns)
    rows = []
    # Truncate to max 500 rows to avoid blowing up context purely for table-QA
    for _, row in df.head(500).iterrows():
        rows.append(" | ".join(str(v).replace("\n", " ").replace("|", "-") for v in row.values))
    return f"| {headers} |\n| {separator} |\n" + "\n".join(f"| {r} |" for r in rows)

def _parse_sheet(df: pd.DataFrame, sheet_name: str) -> Dict[str, Any]:
    rows = [[_clean_cell(v) for v in row] for row in df.values.tolist()]
    layout = _detect_layout(df)
    
    # NEW FAST STRATEGY: Directly cache the Markdown representation of the table
    # This bypasses the 15-minute Python-LLM execution delay during ingestion.
    sheet_markdown = _df_to_markdown_table(df)

    header_idx = _find_header_index(rows)

    # Basic raw rows extraction (we always keep this so Vector search has backup context)
    raw_rows: List[Dict[str, Any]] = []
    for idx, row in enumerate(rows):
        non_empty = _non_empty_values(row)
        if non_empty:
            values = {f"col_{i + 1}": row[i] for i in range(len(row)) if row[i]}
            raw_rows.append({
                "row_id": f"{_to_snake_case(sheet_name)}_raw_{idx}",
                "row_index": idx,
                "section": "",
                "values": values,
                "raw_values": non_empty,
            })

    headers: List[str] = []
    records: List[Dict[str, Any]] = []
    key_values: List[Dict[str, Any]] = []
    raw_rows: List[Dict[str, Any]] = []
    data: List[Dict[str, str]] = []
    current_section = "General"
    prev_label: str = ""  # Fix 2: carry-forward label for orphan rows

    if header_idx is not None:
        raw_header = rows[header_idx]
        # Check if rows immediately above also look like headers (multi-line header support)
        final_headers = list(raw_header)
        for prev_idx in range(max(0, header_idx - 2), header_idx):
            prev_row = rows[prev_idx]
            if _looks_like_header(prev_row) or any("plan" in str(c).lower() for c in prev_row):
                for i in range(min(len(prev_row), len(final_headers))):
                    if prev_row[i] and prev_row[i] != final_headers[i]:
                        final_headers[i] = f"{prev_row[i]} {final_headers[i]}".strip()
        
        headers = _dedupe_headers(final_headers)
        start_idx = header_idx + 1
    else:
        start_idx = 0

    # Preserve every non-empty row for fallback retrieval on unknown layouts.
    # Use actual header names if available; fall back to col_1, col_2 generic names.
    for idx, row in enumerate(rows):
        non_empty = _non_empty_values(row)
        if not non_empty:
            continue
        if headers:
            # Use real header names from detected header row
            values = {}
            for i, cell in enumerate(row):
                if not cell:
                    continue
                key = headers[i] if i < len(headers) else f"col_{i + 1}"
                values[key] = cell
        else:
            values = {f"col_{i + 1}": row[i] for i in range(len(row)) if row[i]}
        raw_rows.append(
            {
                "row_id": f"{_to_snake_case(sheet_name)}_raw_{idx}",
                "row_index": idx,
                "section": "",
                "values": values,
                "raw_values": non_empty,
            }
        )

    for idx in range(start_idx, len(rows)):
        row = rows[idx]
        non_empty = _non_empty_values(row)
        if not non_empty:
            continue

        # Fix 1: Section marker detection - only treat as section if it looks like a label,
        # NOT if it looks like a monetary value or a general data value.
        if len(non_empty) == 1 and len(non_empty[0]) > 1:
            candidate = non_empty[0]
            # Reject values that look like amounts, percentages, yes/no etc.
            is_value_like = bool(
                re.match(r'^[\dRMrm\s,\.%\+\-/]+$', candidate)  # numeric/monetary
                or re.match(r'^(yes|no|covered|not covered|n/a|nil|\-|tbc)$', candidate, re.I)
                or re.search(r'as charged|below rm|above rm|per annum|per visit', candidate, re.I)
            )
            if not is_value_like:
                current_section = candidate
            continue

        values: Dict[str, str] = {}
        for col_idx, cell in enumerate(row):
            if not cell:
                continue
            if headers and col_idx < len(headers):
                key = headers[col_idx]
            else:
                key = f"col_{col_idx + 1}"
            values[key] = cell

        # Fix 2: Carry-forward label for orphan single-value rows.
        # When a row has only the right-side value (col_2), synthesize col_1 from previous row.
        col_1_key = headers[0] if headers else "col_1"
        if col_1_key not in values and len(non_empty) == 1:
            if prev_label:
                values[col_1_key] = prev_label
        elif col_1_key in values:
            prev_label = values[col_1_key]  # remember for next orphan row

        if not values:
            continue

        record = {
            "row_id": f"{_to_snake_case(sheet_name)}_{idx}",
            "row_index": idx,
            "section": current_section,
            "values": values,
            "raw_values": non_empty,
        }
        records.append(record)
        data.append(values)

        if len(non_empty) == 2:
            key_values.append(
                {
                    "row_index": idx,
                    "section": current_section,
                    "key": non_empty[0],
                    "value": non_empty[1],
                }
            )

    # Fix 3: Empty sheet fallback - if no records parsed, dump all cells as raw key-values.
    # This handles cover-page sheets, intro sheets, or unusual layouts.
    if not records and not raw_rows:
        for idx, row in enumerate(rows):
            non_empty = _non_empty_values(row)
            if not non_empty:
                continue
            # Treat 2-column rows as key-value pairs
            if len(non_empty) >= 2:
                records.append({
                    "row_id": f"{_to_snake_case(sheet_name)}_fb_{idx}",
                    "row_index": idx,
                    "section": "Fallback",
                    "values": {"label": non_empty[0], "value": non_empty[1]},
                    "raw_values": non_empty,
                })
                key_values.append({"row_index": idx, "section": "Fallback", "key": non_empty[0], "value": non_empty[1]})
            else:
                # Single-cell rows: store as standalone label
                records.append({
                    "row_id": f"{_to_snake_case(sheet_name)}_fb_{idx}",
                    "row_index": idx,
                    "section": "Fallback",
                    "values": {"label": non_empty[0]},
                    "raw_values": non_empty,
                })

    return {
        "layout": layout,
        "header_row_index": header_idx,
        "headers": headers,
        "records": records,
        "raw_rows": raw_rows,
        "key_values": key_values,
        "data": data,  # Backward-compatible flat records
        "markdown": sheet_markdown, # NEW: The fast-tracked Table-QA string
        "total_records": len(records),
        "total_raw_rows": len(raw_rows),
    }


def _record_to_text(file_name: str, sheet_name: str, record: Dict[str, Any]) -> str:
    values = record.get("values", {})
    value_parts = [f"{k}: {v}" for k, v in values.items()]
    section = record.get("section") or "General"
    return (
        f"File: {file_name} | Sheet: {sheet_name} | Section: {section} | "
        + " | ".join(value_parts)
    )


# ── Ingestion ──────────────────────────────────────────────────────
def ingest_with_intelligence_engine(file_path: str, folder_name: str = "default"):
    file_name = os.path.basename(file_path)
    print(f"\n🚀 [HYBRID-ENGINE] Starting Deterministic Extraction: {file_name}", flush=True)
    start_time_total = time.time()

    try:
        ext = os.path.splitext(file_path)[1].lower()
        sheet_frames = _read_sheet_frames(file_path, ext)
        sheets = list(sheet_frames.keys())
    except Exception as e:
        print(f"[ERROR] Load fail: {e}")
        return {"error": str(e)}

    all_docs: List[Document] = []
    sheet_info: List[Dict[str, Any]] = []
    total_raw_rows = 0
    file_structured_data: Dict[str, Any] = {
        "file_name": file_name,
        "folder": folder_name,
        "ingested_at": datetime.datetime.utcnow().isoformat() + "Z",
        "sheets": {},
    }

    for sheet_name in sheets:
        print(f"\n  🔍 [FILE] Scanning Sheet: {sheet_name}", flush=True)
        try:
            df = sheet_frames.get(sheet_name, pd.DataFrame()).fillna("")
        except Exception as e:
            print(f"  [WARN] Failed reading sheet '{sheet_name}': {e}")
            continue

        sheet_json = _parse_sheet(df, sheet_name)
        file_structured_data["sheets"][sheet_name] = sheet_json
        sheet_info.append(
            {
                "sheet": sheet_name,
                "rows": sheet_json.get("total_records", 0),
                "raw_rows": sheet_json.get("total_raw_rows", 0),
                "layout": sheet_json.get("layout"),
            }
        )
        total_raw_rows += sheet_json.get("total_raw_rows", 0)
        print(f"    LOG: STRUCTURE_DETECTED | Layout: {sheet_json.get('layout')}")

        for record in sheet_json.get("records", []):
            page_text = _record_to_text(file_name, sheet_name, record)
            metadata = {
                "file_name": file_name,
                "source": file_name,
                "sheet": sheet_name,
                "folder": folder_name,
                "layout": sheet_json.get("layout"),
                "record_kind": "table_record",
                "section": record.get("section", "General"),
                "row_id": record.get("row_id"),
                "row_index": record.get("row_index"),
                "record_json": json.dumps(record.get("values", {})),
            }
            all_docs.append(Document(page_content=page_text, metadata=metadata))

        # Add raw rows as additional retrieval surface for unusual layouts.
        for raw_record in sheet_json.get("raw_rows", []):
            raw_text = _record_to_text(file_name, sheet_name, raw_record)
            raw_metadata = {
                "file_name": file_name,
                "source": file_name,
                "sheet": sheet_name,
                "folder": folder_name,
                "layout": sheet_json.get("layout"),
                "record_kind": "raw_row",
                "section": raw_record.get("section", ""),
                "row_id": raw_record.get("row_id"),
                "row_index": raw_record.get("row_index"),
                "record_json": json.dumps(raw_record.get("values", {})),
            }
            all_docs.append(Document(page_content=raw_text, metadata=raw_metadata))

    # Persist cache to memory + disk for deterministic retrieval.
    cache_key = f"{folder_name}::{file_name}"
    EXCEL_CACHE[cache_key] = file_structured_data
    print(f"\n=== [JSON EXTRACTED DATA] ===\n{json.dumps(file_structured_data, indent=2)}\n=============================\n", flush=True)
    structured_json_path = _persist_structured_json(file_name, folder_name, file_structured_data)

    embedding_backend = None
    vector_inserted = False
    vector_insert_error = None
    if all_docs:
        print(f"[QDRANT] Starting Batch Insert | Docs: {len(all_docs)}")
        try:
            embeddings, _, embedding_backend = get_embeddings()
            print(f"[QDRANT] Embedding backend: {embedding_backend}")
            client = get_qdrant_client()
            vector_store = QdrantVectorStore(
                client=client,
                embedding=embeddings,
                collection_name=COLLECTION_NAME,
            )
            # Replace existing vectors for the same folder/file to avoid stale duplicates.
            try:
                client.delete(
                    collection_name=COLLECTION_NAME,
                    points_selector=models.Filter(
                        must=[
                            models.FieldCondition(
                                key="metadata.folder",
                                match=models.MatchValue(value=folder_name),
                            ),
                            models.FieldCondition(
                                key="metadata.file_name",
                                match=models.MatchValue(value=file_name),
                            ),
                        ]
                    ),
                )
            except Exception:
                pass
            vector_store.add_documents(all_docs, batch_size=64)
            vector_inserted = True
        except Exception as e:
            vector_insert_error = str(e)
            print(f"[QDRANT] Vector insert skipped due to embedding/storage error: {e}")

    elapsed = int(time.time() - start_time_total)
    total_rows = sum(x.get("rows", 0) for x in sheet_info)
    print(f"🏆 INGESTION COMPLETE | TIME: {elapsed}s")

    result: Dict[str, Any] = {
        "status": "Ingested (Excel)",
        "num_segments": len(all_docs),
        "sheets_count": len(sheet_info),
        "total_rows": total_rows,
        "total_raw_rows": total_raw_rows,
        "sheet_info": sheet_info,
        "structured_json": structured_json_path,
        "embedding_backend": embedding_backend,
        "vectors_inserted": vector_inserted,
    }
    if vector_insert_error:
        result["vector_error"] = vector_insert_error
    return result
